#!/usr/bin/env python3
"""Asset grouping runner for Office Shield.

Reads the Excel template, ensures terminal groups exist, queries terminals by MAC,
and optionally moves matched terminal GUIDs to target groups.
"""

from __future__ import annotations

import argparse
import getpass
import json
import os
import ssl
import sys
import time
import urllib.error
import urllib.request
import warnings
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DEFAULT_WORKDIR = Path("D:/tmp/asset-grouping-work")
DEFAULT_TEMPLATE = DEFAULT_WORKDIR / "资产分组导入模板.xlsx"
DEFAULT_CONFIG = DEFAULT_WORKDIR / "asset_grouping_config.json"


DEFAULT_CONFIG_DATA = {
    "template_path": str(DEFAULT_TEMPLATE),
    "output_dir": str(DEFAULT_WORKDIR),
    "api_base_url": "https://172.21.193.214:53443",
    "internal_api_base_url": "https://172.21.193.214:53443",
    "external_api_base_url": "https://192.168.224.84:53443",
    "query_scope_path": "全部终端",
    "internal_cookie": "",
    "external_cookie": "",
    "default_network": "内网",
    "strict_tls": False,
    "keep_history": False,
}

NETWORK_ORDER = ["internal", "external"]
NETWORK_LABELS = {
    "internal": "内网",
    "external": "外网",
}

SAFE_EXECUTE_BATCH_SIZE = 1


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="按照资产分组导入模板自动建组、查终端、移动终端。",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument("--config", default=str(DEFAULT_CONFIG), help="配置文件路径")
    parser.add_argument("--init-config", action="store_true", help="创建/补全配置文件后退出")
    parser.add_argument("--interactive", "-i", action="store_true", help="启动终端交互界面")
    parser.add_argument("--run", action="store_true", help="不进入交互界面，按命令行参数直接运行")
    parser.add_argument("--template", default=None, help="模板路径")
    parser.add_argument("--output-dir", default=None, help="报告输出目录")
    parser.add_argument("--cookie", default=None, help="平台 Cookie")
    parser.add_argument("--base-url", default=None, help="覆盖模板中的 api_base_url")
    parser.add_argument("--internal-cookie", default=None, help="内网平台 Cookie")
    parser.add_argument("--external-cookie", default=None, help="外网平台 Cookie")
    parser.add_argument("--internal-base-url", default=None, help="内网平台 API 地址")
    parser.add_argument("--external-base-url", default=None, help="外网平台 API 地址")
    parser.add_argument("--default-network", default=None, help="终端清单网络区域为空/无法识别时的默认网络：内网 或 外网")
    parser.add_argument("--query-scope-path", default=None, help="查询终端的范围分组路径")
    parser.add_argument("--compare-export", default=None, help="比对平台实时分组与导出的分组 Excel")
    parser.add_argument("--export-invalid-mac", nargs="?", const="", default=None, help="导出终端清单中不合法的 MAC；可选输出 xlsx 路径")
    parser.add_argument("--export-not-found-terminals", nargs="?", const="", default=None, help="从最近报告导出平台未查询到的终端；可选输出 xlsx 路径")
    parser.add_argument("--execute", action="store_true", help="在模板 dry_run=FALSE 时真正执行建组和移动")
    parser.add_argument("--local-only", action="store_true", help="只解析模板，不访问平台")
    parser.add_argument("--strict-tls", action="store_true", help="启用严格 TLS 校验；默认兼容内网自签证书")
    parser.add_argument("--keep-history", action="store_true", help="同时保留带时间戳的历史报告；默认只更新 latest 报告")
    return parser.parse_args()


def load_config(config_path: Path) -> dict[str, Any]:
    if not config_path.exists():
        return dict(DEFAULT_CONFIG_DATA)
    try:
        loaded = json.loads(config_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"配置文件不是合法 JSON: {config_path} ({exc})") from exc
    if not isinstance(loaded, dict):
        raise RuntimeError(f"配置文件顶层必须是 JSON 对象: {config_path}")
    config = dict(DEFAULT_CONFIG_DATA)
    config.update(loaded)
    legacy_cookie = text(loaded.get("cookie"))
    if legacy_cookie:
        if not text(config.get("internal_cookie")):
            config["internal_cookie"] = legacy_cookie
    config.pop("cookie", None)
    return config


def save_config(config_path: Path, config: dict[str, Any]) -> None:
    config_path.parent.mkdir(parents=True, exist_ok=True)
    ordered = dict(DEFAULT_CONFIG_DATA)
    ordered.update(config)
    ordered.pop("cookie", None)
    config_path.write_text(json.dumps(ordered, ensure_ascii=False, indent=2), encoding="utf-8")


def apply_config(args: argparse.Namespace, config: dict[str, Any]) -> argparse.Namespace:
    cli_cookie = args.cookie
    env_cookie = os.getenv("UES_COOKIE") or os.getenv("OFFICE_SHIELD_COOKIE") or ""
    env_internal_cookie = os.getenv("UES_INTERNAL_COOKIE") or os.getenv("OFFICE_SHIELD_INTERNAL_COOKIE") or env_cookie
    env_external_cookie = os.getenv("UES_EXTERNAL_COOKIE") or os.getenv("OFFICE_SHIELD_EXTERNAL_COOKIE") or env_cookie
    common_cookie = text(config.get("cookie"))
    internal_cookie = text(config.get("internal_cookie")) or common_cookie
    external_cookie = text(config.get("external_cookie")) or common_cookie
    legacy_base_url = text(config.get("api_base_url")) or DEFAULT_CONFIG_DATA["api_base_url"]
    args.template = args.template or text(config.get("template_path")) or str(DEFAULT_TEMPLATE)
    args.output_dir = args.output_dir or text(config.get("output_dir")) or str(DEFAULT_WORKDIR)
    args.cookie = cli_cookie if cli_cookie is not None else env_cookie or common_cookie
    args.internal_cookie = args.internal_cookie if args.internal_cookie is not None else cli_cookie or env_internal_cookie or internal_cookie
    args.external_cookie = args.external_cookie if args.external_cookie is not None else cli_cookie or env_external_cookie or external_cookie
    args.internal_base_url = args.internal_base_url or args.base_url or text(config.get("internal_api_base_url")) or legacy_base_url
    args.external_base_url = args.external_base_url or text(config.get("external_api_base_url")) or DEFAULT_CONFIG_DATA["external_api_base_url"]
    args.base_url = args.base_url if args.base_url is not None else args.internal_base_url
    args.default_network = args.default_network or text(config.get("default_network")) or "内网"
    args.query_scope_path = args.query_scope_path or text(config.get("query_scope_path")) or "全部终端"
    args.strict_tls = bool(args.strict_tls or config.get("strict_tls"))
    args.keep_history = bool(args.keep_history or config.get("keep_history"))
    return args


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).strip()


def enabled(value: Any) -> bool:
    return text(value).upper() == "Y"


def normalize_path(path: Any, separator: str = ">") -> str:
    return separator.join(part for part in (text(item) for item in text(path).split(separator)) if part)


def split_path(path: str, separator: str = ">") -> list[str]:
    return [part for part in normalize_path(path, separator).split(separator) if part]


def is_full_group_path(path: Any, separator: str = ">") -> bool:
    normalized = normalize_path(path, separator)
    return normalized == "全部终端" or normalized.startswith(f"全部终端{separator}")


def normalize_mac(value: Any) -> str:
    raw = text(value).upper()
    separators = ":：-－–—. _"
    for separator in separators:
        raw = raw.replace(separator, "")
    return raw


def format_mac_for_api(value: Any) -> str:
    normalized = normalize_mac(value)
    if len(normalized) == 12:
        return ":".join(normalized[i : i + 2] for i in range(0, 12, 2))
    return text(value).upper()


def valid_normalized_mac(value: str) -> bool:
    return len(value) == 12 and all(char in "0123456789ABCDEF" for char in value)


def invalid_mac_chars(value: str) -> str:
    chars = sorted({char for char in value if char not in "0123456789ABCDEF"})
    return ",".join(chars)


def normalize_network_key(value: Any, default_key: str = "internal") -> str:
    raw = text(value).lower()
    if raw in {"external", "outer", "wan", "internet"} or "外网" in raw or "公网" in raw or "互联网" in raw:
        return "external"
    if raw in {"internal", "inner", "lan"} or "内网" in raw or "局域网" in raw:
        return "internal"
    if text(value) == "外":
        return "external"
    if text(value) == "内":
        return "internal"
    return default_key or "internal"


def network_label(network_key: str) -> str:
    return NETWORK_LABELS.get(network_key, network_key or "未知")


def expand_group_paths(paths: list[str], separator: str = ">") -> list[str]:
    expanded: set[str] = set()
    for path in paths:
        parts = split_path(path, separator)
        for index in range(1, len(parts) + 1):
            expanded.add(separator.join(parts[:index]))
    return list(expanded)


def chunks(items: list[str], size: int) -> list[list[str]]:
    return [items[index : index + size] for index in range(0, len(items), size)]


def clipped(value: Any, limit: int = 36) -> str:
    content = text(value)
    if len(content) <= limit:
        return content
    return f"{content[: max(limit - 3, 0)]}..."


class ProgressBar:
    def __init__(self, total: int, label: str, enabled: bool = True) -> None:
        self.total = max(int(total or 0), 0)
        self.label = label
        self.enabled = bool(enabled and self.total)
        self.current = 0
        self.started = time.monotonic()
        self.last_len = 0
        if self.enabled:
            self._render("")

    def update(self, step: int = 1, detail: str = "") -> None:
        if not self.enabled:
            return
        self.current = min(self.total, self.current + step)
        self._render(detail)

    def finish(self, detail: str = "完成") -> None:
        if not self.enabled:
            return
        self.current = self.total
        self._render(detail)
        sys.stdout.write("\n")
        sys.stdout.flush()
        self.enabled = False

    def close(self, detail: str = "") -> None:
        if not self.enabled:
            return
        self._render(detail)
        sys.stdout.write("\n")
        sys.stdout.flush()
        self.enabled = False

    def _render(self, detail: str) -> None:
        width = 28
        ratio = self.current / self.total if self.total else 1
        filled = min(width, int(width * ratio))
        bar = "#" * filled + "-" * (width - filled)
        percent = ratio * 100
        elapsed = int(time.monotonic() - self.started)
        detail_text = clipped(detail, 34)
        line = f"\r{self.label} [{bar}] {self.current}/{self.total} {percent:5.1f}% {elapsed:>4}s"
        if detail_text:
            line += f" {detail_text}"
        padding = " " * max(self.last_len - len(line), 0)
        sys.stdout.write(line + padding)
        sys.stdout.flush()
        self.last_len = len(line)


def row_dicts(ws: Any, header_row: int, max_row: int, max_col: int) -> list[dict[str, Any]]:
    header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, min_col=1, max_col=max_col, values_only=True))
    headers = [text(value) for value in header_values]
    rows: list[dict[str, Any]] = []
    for excel_row, values in enumerate(
        ws.iter_rows(min_row=header_row + 1, max_row=max_row, min_col=1, max_col=max_col, values_only=True),
        start=header_row + 1,
    ):
        record: dict[str, Any] = {"__rowNumber": excel_row}
        for index, header in enumerate(headers):
            record[header] = values[index] if index < len(values) else None
        rows.append(record)
    return rows


def build_target_group_usage(move_plan: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_path: dict[str, dict[str, Any]] = {}
    for item in move_plan:
        path = item["targetPath"]
        current = by_path.setdefault(
            path,
            {
                "targetPath": path,
                "total": 0,
                "byNetwork": {key: 0 for key in NETWORK_ORDER},
                "bySource": {},
                "sampleRows": [],
            },
        )
        current["total"] += 1
        network_key = item.get("networkKey") or "internal"
        current["byNetwork"][network_key] = current["byNetwork"].get(network_key, 0) + 1
        source = item.get("targetSource") or "unknown"
        current["bySource"][source] = current["bySource"].get(source, 0) + 1
        if len(current["sampleRows"]) < 5:
            current["sampleRows"].append(item.get("sourceRow"))
    return sorted(by_path.values(), key=lambda row: row["targetPath"])


def load_template_plan(template_path: Path, default_network: str = "internal") -> dict[str, Any]:
    workbook = load_workbook(template_path, read_only=True, data_only=True)

    def config_map() -> dict[str, str]:
        ws = workbook["执行配置"]
        config: dict[str, str] = {}
        for key, value in ws.iter_rows(min_row=7, max_row=50, min_col=1, max_col=2, values_only=True):
            key_text = text(key)
            if key_text:
                config[key_text] = text(value)
        return config

    config = config_map()
    separator = config.get("path_separator") or ">"
    blank_dept_literal = config.get("blank_dept_literal") or "(空白)"
    unmatched_path = normalize_path(config.get("unmatched_path") or "全部终端>门急诊>待确认责任科室", separator)
    default_network_key = normalize_network_key(config.get("default_network") or default_network, "internal")
    create_declared_groups_without_terminals = text(config.get("create_declared_groups_without_terminals")).upper() == "TRUE"

    group_rows = []
    for row in row_dicts(workbook["分组树"], 6, 500, 7):
        if not enabled(row.get("启用")):
            continue
        path = normalize_path(row.get("完整分组路径"), separator)
        if not path:
            continue
        group_rows.append(
            {
                "sourceRow": row["__rowNumber"],
                "path": path,
                "description": text(row.get("分组描述")),
                "allowCreate": text(row.get("是否允许新建")).upper() != "N",
            }
        )

    mapping_rows = []
    for row in row_dicts(workbook["科室映射"], 6, 1000, 8):
        if not enabled(row.get("启用")):
            continue
        source_dept = text(row.get("来源设备责任科室")) or blank_dept_literal
        target_path = normalize_path(row.get("目标分组路径"), separator)
        if not source_dept or not target_path:
            continue
        try:
            priority = int(float(text(row.get("优先级")) or "0"))
        except ValueError:
            priority = 0
        mapping_rows.append(
            {
                "sourceRow": row["__rowNumber"],
                "sourceDept": source_dept,
                "targetPath": target_path,
                "matchType": text(row.get("匹配方式")) or "exact",
                "priority": priority,
                "allowFallback": text(row.get("是否允许落入兜底")).upper() == "Y",
            }
        )
    mapping_rows.sort(key=lambda item: item["priority"], reverse=True)

    def find_mapping(dept: str) -> dict[str, Any] | None:
        source = text(dept) or blank_dept_literal
        for rule in mapping_rows:
            match_type = rule["matchType"]
            pattern = rule["sourceDept"]
            if match_type == "exact" and source == pattern:
                return rule
            if match_type == "contains" and pattern in source:
                return rule
            if match_type == "regex":
                import re

                try:
                    if re.search(pattern, source):
                        return rule
                except re.error:
                    return None
        return None

    terminal_rows = []
    for row in row_dicts(workbook["终端清单"], 6, 5000, 15):
        if not enabled(row.get("启用")):
            continue
        network_area = text(row.get("网络区域"))
        network_key = normalize_network_key(network_area, default_network_key)
        terminal_rows.append(
            {
                "sourceRow": row["__rowNumber"],
                "assetCode": text(row.get("资产编号")),
                "mac": text(row.get("MAC地址")),
                "normalizedMac": normalize_mac(row.get("MAC地址")),
                "hostname": text(row.get("主机名")),
                "dept": text(row.get("设备责任科室")),
                "manualTargetPath": normalize_path(row.get("手动目标分组路径"), separator),
                "campus": text(row.get("院区")),
                "building": text(row.get("楼栋")),
                "floor": text(row.get("楼层")),
                "area": text(row.get("分区")),
                "location": text(row.get("具体物理地址")),
                "networkArea": network_area,
                "networkKey": network_key,
                "networkInferred": network_key == default_network_key
                and network_area
                and normalize_network_key(network_area, "__unknown__") == "__unknown__",
                "networkEmpty": not network_area,
            }
        )

    target_paths: set[str] = set()
    if create_declared_groups_without_terminals:
        target_paths.update(row["path"] for row in group_rows)
        target_paths.update(rule["targetPath"] for rule in mapping_rows)

    validation: list[dict[str, Any]] = []
    move_plan: list[dict[str, Any]] = []

    for terminal in terminal_rows:
        if terminal["networkEmpty"]:
            validation.append(
                {
                    "source": "终端清单",
                    "row": terminal["sourceRow"],
                    "severity": "warning",
                    "field": "网络区域",
                    "value": "",
                    "message": f"网络区域为空，已按默认网络 {network_label(default_network_key)} 处理。",
                }
            )
        elif terminal["networkInferred"]:
            validation.append(
                {
                    "source": "终端清单",
                    "row": terminal["sourceRow"],
                    "severity": "warning",
                    "field": "网络区域",
                    "value": terminal["networkArea"],
                    "message": f"网络区域无法识别，已按默认网络 {network_label(default_network_key)} 处理。",
                }
            )

        mapping = None
        if terminal["manualTargetPath"]:
            if not is_full_group_path(terminal["manualTargetPath"], separator):
                validation.append(
                    {
                        "source": "终端清单",
                        "row": terminal["sourceRow"],
                        "severity": "error",
                        "field": "手动目标分组路径",
                        "value": terminal["manualTargetPath"],
                        "message": "手动目标分组路径必须填写完整路径，例如 全部终端>门急诊>门诊>门诊部01护理单元。",
                    }
                )
                continue
            target_path = terminal["manualTargetPath"]
            target_source = "manual"
        else:
            mapping = find_mapping(terminal["dept"])
            if mapping:
                target_path = mapping["targetPath"]
                target_source = "mapping"
            else:
                target_path = unmatched_path
                target_source = "unmatched_fallback"
                validation.append(
                    {
                        "source": "终端清单",
                        "row": terminal["sourceRow"],
                        "severity": "warning",
                        "field": "设备责任科室",
                        "value": terminal["dept"] or blank_dept_literal,
                        "message": "未找到科室映射，已计划进入兜底分组。",
                    }
                )

        if not terminal["normalizedMac"]:
            validation.append(
                {
                    "source": "终端清单",
                    "row": terminal["sourceRow"],
                    "severity": "error",
                    "field": "MAC地址",
                    "value": terminal["mac"],
                    "message": "当前执行器按 MAC 调用平台查询接口；请填写 MAC 地址。",
                }
            )
            continue
        if not valid_normalized_mac(terminal["normalizedMac"]):
            validation.append(
                {
                    "source": "终端清单",
                    "row": terminal["sourceRow"],
                    "severity": "error",
                    "field": "MAC地址",
                    "value": terminal["mac"],
                    "message": "MAC 地址需能规范化为 12 位十六进制；支持 AA-BB-CC-DD-EE-FF、AA:BB:CC:DD:EE:FF、AABBCCDDEEFF 等格式。",
                }
            )
            continue

        target_paths.add(target_path)
        move_plan.append(
            {
                "sourceRow": terminal["sourceRow"],
                "assetCode": terminal["assetCode"],
                "mac": terminal["mac"],
                "apiMac": format_mac_for_api(terminal["mac"]),
                "normalizedMac": terminal["normalizedMac"],
                "hostname": terminal["hostname"],
                "dept": terminal["dept"] or blank_dept_literal,
                "networkArea": terminal["networkArea"] or network_label(terminal["networkKey"]),
                "networkKey": terminal["networkKey"],
                "networkName": network_label(terminal["networkKey"]),
                "targetPath": target_path,
                "targetSource": target_source,
                "mappingSourceRow": mapping["sourceRow"] if mapping else None,
                "location": " / ".join(
                    item
                    for item in [terminal["campus"], terminal["building"], terminal["floor"], terminal["area"], terminal["location"]]
                    if item
                ),
            }
        )

    allow_create_by_path = {row["path"]: row["allowCreate"] for row in group_rows}
    description_by_path = {row["path"]: row["description"] for row in group_rows}
    expanded_paths = expand_group_paths(list(target_paths), separator)
    expanded_paths.sort(key=lambda path: (len(split_path(path, separator)), path))
    used_group_paths = set(expanded_paths)
    unused_group_rows = [row for row in group_rows if row["path"] not in used_group_paths]

    create_group_plan = []
    for path in expanded_paths:
        if not is_full_group_path(path, separator):
            continue
        parts = split_path(path, separator)
        create_group_plan.append(
            {
                "path": path,
                "name": parts[-1],
                "parentPath": separator.join(parts[:-1]),
                "description": description_by_path.get(path, ""),
                "allowCreate": allow_create_by_path.get(path, True),
                "declaredInTemplate": any(row["path"] == path for row in group_rows),
            }
        )
    target_group_usage = build_target_group_usage(move_plan)

    return {
        "templatePath": str(template_path),
        "config": {
            "apiBaseUrl": config.get("api_base_url"),
            "dryRun": text(config.get("dry_run")).upper() != "FALSE",
            "createMissingGroups": text(config.get("create_missing_groups")).upper() != "FALSE",
            "moveTerminals": text(config.get("move_terminals")).upper() != "FALSE",
            "deleteExtraGroups": text(config.get("delete_extra_groups")).upper() == "TRUE",
            "terminalQueryPageSize": int(float(text(config.get("terminal_query_page_size")) or "200")),
            "batchSize": max(1, int(float(text(config.get("batch_size")) or str(SAFE_EXECUTE_BATCH_SIZE)))),
            "executeBatchSize": SAFE_EXECUTE_BATCH_SIZE,
            "moveVerifyRetries": max(0, int(float(text(config.get("move_verify_retries")) or "2"))),
            "moveVerifyDelaySeconds": max(0.0, float(text(config.get("move_verify_delay_seconds")) or "0.8")),
            "createDeclaredGroupsWithoutTerminals": create_declared_groups_without_terminals,
            "rootPath": normalize_path(config.get("root_path") or "全部终端", separator),
            "separator": separator,
            "multiTerminalMatchPolicy": "move_all_matches_to_target_group",
        },
        "counts": {
            "enabledGroupRows": len(group_rows),
            "unusedEnabledGroupRows": len(unused_group_rows),
            "enabledMappingRows": len(mapping_rows),
            "enabledTerminalRows": len(terminal_rows),
            "plannedTargetGroups": len(target_group_usage),
            "plannedGroups": len(create_group_plan),
            "plannedTerminalMoves": len(move_plan),
            "plannedTerminalMovesByNetwork": {
                key: sum(1 for item in move_plan if item.get("networkKey") == key)
                for key in NETWORK_ORDER
            },
            "validationErrors": sum(1 for item in validation if item["severity"] == "error"),
            "validationWarnings": sum(1 for item in validation if item["severity"] == "warning"),
        },
        "createGroupPlan": create_group_plan,
        "targetGroupUsage": target_group_usage,
        "movePlan": move_plan,
        "validation": validation,
    }


class ApiClient:
    def __init__(self, base_url: str, cookie: str, insecure: bool) -> None:
        self.base_url = base_url.rstrip("/")
        self.cookie = cookie
        self.context = ssl._create_unverified_context() if insecure else None

    def request(self, method: str, path: str, body: Any | None = None) -> Any:
        url = f"{self.base_url}{path}"
        payload = json.dumps(body, ensure_ascii=False).encode("utf-8") if body is not None else None
        headers = {
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "zh-CN",
            "Accept-Encoding": "identity",
            "Referer": f"{self.base_url}/ues/base/terminals",
            "Cookie": self.cookie,
        }
        if payload is not None:
            headers["Content-Type"] = "application/json"
        request = urllib.request.Request(url, data=payload, headers=headers, method=method)
        try:
            with urllib.request.urlopen(request, context=self.context, timeout=60) as response:
                raw = response.read().decode("utf-8")
        except urllib.error.HTTPError as error:
            raw = error.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"{method} {path} HTTP {error.code}: {raw[:500]}") from error
        except urllib.error.URLError as error:
            raise RuntimeError(f"{method} {path} 请求失败: {error}") from error
        try:
            return json.loads(raw) if raw else None
        except json.JSONDecodeError as error:
            raise RuntimeError(f"{method} {path} 返回非 JSON: {raw[:500]}") from error

    def get_groups(self) -> list[dict[str, Any]]:
        data = self.request("GET", "/groups")
        assert_api_ok(data, "GET /groups")
        return data.get("data") or []

    def create_group(self, name: str, parent_guid: str, description: str = "") -> dict[str, Any]:
        data = self.request(
            "POST",
            "/groups",
            {
                "description": description or "",
                "lockChildGroupPolicy": False,
                "name": name,
                "parentGuid": parent_guid,
                "policyType": "",
                "autoGroupRule": {
                    "ruleItemsBatch": [],
                    "reGroup": 0,
                    "syncBind": False,
                },
            },
        )
        assert_api_ok(data, f"POST /groups {name}")
        return data.get("data") or {}

    def query_terminals_by_mac(self, scope_guid: str, mac: str, page_size: int) -> list[dict[str, Any]]:
        terminals: list[dict[str, Any]] = []
        cur_page = 1
        while True:
            body = {
                "sort": "",
                "groupGuidList": [scope_guid],
                "subProductCode": "DAS-UES-SMP",
                "includeChild": True,
                "dumbTerminal": 0,
                "mac": mac,
            }
            data = self.request("POST", f"/terminals/query?curPage={cur_page}&pageSize={page_size}", body)
            assert_api_ok(data, f"POST /terminals/query mac={mac}")
            page_data = data.get("data") or {}
            rows = page_data.get("list") or []
            terminals.extend(rows)
            total_row = int(page_data.get("totalRow") or len(rows))
            if not rows or cur_page * page_size >= total_row:
                break
            cur_page += 1
        return terminals

    def move_terminals(self, terminal_guid_list: list[str], target_group_guid: str) -> Any:
        data = self.request(
            "POST",
            "/terminals/move",
            {
                "terminalGuidList": terminal_guid_list,
                "groupGuidList": [],
                "targetGroupGuid": target_group_guid,
            },
        )
        assert_move_ok(data, "POST /terminals/move")
        return data


def assert_api_ok(data: Any, label: str) -> None:
    if not isinstance(data, dict) or data.get("code") != 0:
        raise RuntimeError(f"{label} 失败: {json.dumps(data, ensure_ascii=False)[:1000]}")


def assert_move_ok(data: Any, label: str) -> None:
    if isinstance(data, list):
        failed = [item for item in data if not (isinstance(item, dict) and (item.get("code") == 0 or item.get("success") is True))]
        if failed:
            raise RuntimeError(f"{label} 部分失败: {json.dumps(failed, ensure_ascii=False)[:1000]}")
        return
    assert_api_ok(data, label)


def flatten_groups(nodes: list[dict[str, Any]], separator: str = ">", parent_path: str = "") -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for node in nodes or []:
        name = text(node.get("name"))
        if not name:
            continue
        path = f"{parent_path}{separator}{name}" if parent_path else name
        rows.append(
            {
                "guid": text(node.get("guid")),
                "name": name,
                "parentGuid": text(node.get("parentGuid")),
                "path": path,
                "raw": node,
            }
        )
        rows.extend(flatten_groups(node.get("children") or [], separator, path))
    return rows


def build_path_index(group_tree: list[dict[str, Any]], separator: str) -> dict[str, Any]:
    flattened = flatten_groups(group_tree, separator)
    by_path: dict[str, dict[str, Any]] = {}
    duplicates: list[str] = []
    for group in flattened:
        if group["path"] in by_path:
            duplicates.append(group["path"])
        by_path[group["path"]] = group
    return {"flattened": flattened, "byPath": by_path, "duplicates": duplicates}


def platform_configs_from_args(args: argparse.Namespace) -> dict[str, dict[str, Any]]:
    return {
        "internal": {
            "networkKey": "internal",
            "networkName": "内网",
            "apiBaseUrl": text(getattr(args, "internal_base_url", "")) or text(getattr(args, "base_url", "")),
            "cookie": text(getattr(args, "internal_cookie", "")) or text(getattr(args, "cookie", "")),
        },
        "external": {
            "networkKey": "external",
            "networkName": "外网",
            "apiBaseUrl": text(getattr(args, "external_base_url", "")),
            "cookie": text(getattr(args, "external_cookie", "")) or text(getattr(args, "cookie", "")),
        },
    }


def planned_network_keys(plan: dict[str, Any]) -> list[str]:
    keys = {item.get("networkKey") or "internal" for item in plan.get("movePlan") or []}
    return [key for key in NETWORK_ORDER if key in keys] + sorted(keys - set(NETWORK_ORDER))


def filter_plan_for_network(plan: dict[str, Any], network_key: str) -> dict[str, Any]:
    separator = plan["config"]["separator"]
    move_plan = [dict(item) for item in plan.get("movePlan") or [] if (item.get("networkKey") or "internal") == network_key]
    needed_paths = set(expand_group_paths([item["targetPath"] for item in move_plan], separator))
    create_group_plan = [dict(item) for item in plan.get("createGroupPlan") or [] if item["path"] in needed_paths]
    network_plan = dict(plan)
    network_plan["movePlan"] = move_plan
    network_plan["createGroupPlan"] = create_group_plan
    network_plan["targetGroupUsage"] = build_target_group_usage(move_plan)
    network_plan["counts"] = dict(plan.get("counts") or {})
    network_plan["counts"]["plannedGroups"] = len(create_group_plan)
    network_plan["counts"]["plannedTargetGroups"] = len(network_plan["targetGroupUsage"])
    network_plan["counts"]["plannedTerminalMoves"] = len(move_plan)
    return network_plan


def new_platform_report(network_key: str, platform_config: dict[str, Any]) -> dict[str, Any]:
    return {
        "networkKey": network_key,
        "networkName": platform_config["networkName"],
        "apiBaseUrl": platform_config["apiBaseUrl"],
        "status": "pending",
        "platform": {
            "networkKey": network_key,
            "networkName": platform_config["networkName"],
            "apiBaseUrl": platform_config["apiBaseUrl"],
        },
        "groups": [],
        "terminalQueries": [],
        "moves": {"resolvedTerminalGuidCount": 0, "conflicts": [], "batches": [], "verifications": []},
        "validation": [],
    }


def tag_platform_records(platform_report: dict[str, Any]) -> None:
    network_key = platform_report["networkKey"]
    network_name = platform_report["networkName"]
    for key in ["groups", "terminalQueries"]:
        for item in platform_report.get(key) or []:
            item.setdefault("networkKey", network_key)
            item.setdefault("networkName", network_name)
    moves = platform_report.get("moves") or {}
    for item in moves.get("conflicts") or []:
        item.setdefault("networkKey", network_key)
        item.setdefault("networkName", network_name)
    for item in moves.get("batches") or []:
        item.setdefault("networkKey", network_key)
        item.setdefault("networkName", network_name)
    for item in moves.get("verifications") or []:
        item.setdefault("networkKey", network_key)
        item.setdefault("networkName", network_name)
    for item in platform_report.get("validation") or []:
        item.setdefault("networkKey", network_key)
        item.setdefault("networkName", network_name)


def merge_platform_report(report: dict[str, Any], platform_report: dict[str, Any]) -> None:
    tag_platform_records(platform_report)
    report["platformRuns"].append(platform_report)
    report["groups"].extend(platform_report.get("groups") or [])
    report["terminalQueries"].extend(platform_report.get("terminalQueries") or [])
    report["validation"].extend(platform_report.get("validation") or [])
    report["moves"]["resolvedTerminalGuidCount"] += (platform_report.get("moves") or {}).get("resolvedTerminalGuidCount", 0)
    report["moves"]["conflicts"].extend((platform_report.get("moves") or {}).get("conflicts") or [])
    report["moves"]["batches"].extend((platform_report.get("moves") or {}).get("batches") or [])
    report["moves"].setdefault("verifications", []).extend((platform_report.get("moves") or {}).get("verifications") or [])


def ensure_groups(client: ApiClient, plan: dict[str, Any], execute: bool, report: dict[str, Any], progress_enabled: bool = True) -> dict[str, Any]:
    progress = ProgressBar(len(plan["createGroupPlan"]) + 1, f"[{report.get('networkName', '')}] 分组阶段", progress_enabled)
    try:
        group_tree = client.get_groups()
        progress.update(detail="已读取 /groups")
        index = build_path_index(group_tree, plan["config"]["separator"])
        report["platform"]["groupCount"] = len(index["flattened"])
        report["platform"]["duplicateGroupPaths"] = index["duplicates"]

        for group in plan["createGroupPlan"]:
            try:
                if group["path"] == "全部终端":
                    current = index["byPath"].get(group["path"])
                    report["groups"].append({"action": "root", "status": "exists" if current else "missing", "path": group["path"], "guid": current.get("guid") if current else ""})
                    continue

                existing = index["byPath"].get(group["path"])
                if existing:
                    report["groups"].append({"action": "ensure_group", "status": "exists", "path": group["path"], "guid": existing["guid"]})
                    continue

                if not group["allowCreate"] or not plan["config"]["createMissingGroups"]:
                    report["groups"].append({"action": "ensure_group", "status": "blocked", "path": group["path"], "reason": "模板或配置不允许自动新建"})
                    report["validation"].append({"severity": "error", "source": "分组树", "field": "完整分组路径", "value": group["path"], "message": "目标分组不存在，且不允许自动新建。"})
                    continue

                parent = index["byPath"].get(group["parentPath"])
                if not parent:
                    report["groups"].append({"action": "ensure_group", "status": "blocked", "path": group["path"], "parentPath": group["parentPath"], "reason": "父分组不存在"})
                    report["validation"].append({"severity": "error", "source": "分组树", "field": "完整分组路径", "value": group["path"], "message": f"父分组不存在: {group['parentPath']}"})
                    continue

                if not execute:
                    report["groups"].append({"action": "create_group", "status": "dry_run", "path": group["path"], "name": group["name"], "parentPath": group["parentPath"], "parentGuid": parent["guid"]})
                    continue

                created = client.create_group(group["name"], parent["guid"], group.get("description", ""))
                created_group = {
                    "guid": text(created.get("guid")),
                    "name": text(created.get("name")) or group["name"],
                    "parentGuid": text(created.get("parentGuid")) or parent["guid"],
                    "path": group["path"],
                    "raw": created,
                }
                index["byPath"][group["path"]] = created_group
                index["flattened"].append(created_group)
                report["groups"].append({"action": "create_group", "status": "created", "path": group["path"], "guid": created_group["guid"], "parentGuid": parent["guid"]})
            finally:
                progress.update(detail=group.get("path", ""))
    except Exception:
        progress.close("失败")
        raise
    progress.finish()

    return index


def resolve_terminal_moves(client: ApiClient, plan: dict[str, Any], group_index: dict[str, Any], query_scope_path: str, report: dict[str, Any], progress_enabled: bool = True) -> list[dict[str, Any]]:
    separator = plan["config"]["separator"]
    scope_path = normalize_path(query_scope_path or "全部终端", separator)
    fallback_scope_path = normalize_path(plan["config"]["rootPath"] or "全部终端", separator)
    scope = group_index["byPath"].get(scope_path) or group_index["byPath"].get(fallback_scope_path)
    if not scope:
        report["validation"].append({"severity": "error", "source": "执行参数", "field": "queryScopePath", "value": scope_path, "message": f"查询终端的范围分组不存在，也未找到备用范围 {fallback_scope_path}"})
        return []

    report["platform"]["queryScopePath"] = scope["path"]
    report["platform"]["queryScopeGuid"] = scope["guid"]
    cache_by_mac: dict[str, list[dict[str, Any]]] = {}
    resolved: list[dict[str, Any]] = []

    progress = ProgressBar(len(plan["movePlan"]), f"[{report.get('networkName', '')}] 查询终端", progress_enabled)
    try:
        for item in plan["movePlan"]:
            try:
                target = group_index["byPath"].get(item["targetPath"])
                if not target:
                    report["terminalQueries"].append({"sourceRow": item["sourceRow"], "mac": item["mac"], "status": "blocked", "reason": f"目标分组不存在: {item['targetPath']}"})
                    report["validation"].append({"severity": "error", "source": "终端清单", "row": item["sourceRow"], "field": "目标分组路径", "value": item["targetPath"], "message": "目标分组不存在，无法移动终端。"})
                    continue

                matches = cache_by_mac.get(item["normalizedMac"])
                if matches is None:
                    matches = client.query_terminals_by_mac(scope["guid"], item["apiMac"], plan["config"]["terminalQueryPageSize"])
                    cache_by_mac[item["normalizedMac"]] = matches

                matched_guids = [text(row.get("guid")) for row in matches if text(row.get("guid"))]
                report["terminalQueries"].append(
                    {
                        "sourceRow": item["sourceRow"],
                        "mac": item["mac"],
                        "apiMac": item["apiMac"],
                        "targetPath": item["targetPath"],
                        "targetGuid": target["guid"],
                        "queryScopePath": scope["path"],
                        "queryScopeGuid": scope["guid"],
                        "matchCount": len(matches),
                        "matchedGuids": matched_guids,
                        "status": "matched" if matches else "not_found",
                    }
                )

                if not matches:
                    report["validation"].append({"severity": "warning", "source": "终端清单", "row": item["sourceRow"], "field": "MAC地址", "value": item["mac"], "message": "平台未按 MAC 查询到终端。"})
                    continue

                for terminal in matches:
                    terminal_guid = text(terminal.get("guid"))
                    if not terminal_guid:
                        continue
                    resolved.append(
                        {
                            "sourceRow": item["sourceRow"],
                            "mac": item["mac"],
                            "apiMac": item["apiMac"],
                            "normalizedMac": item["normalizedMac"],
                            "terminalGuid": terminal_guid,
                            "currentGroupGuid": text(terminal.get("groupGuid")),
                            "terminalName": text(terminal.get("terminalName")),
                            "hostName": text(terminal.get("hostName")),
                            "targetPath": item["targetPath"],
                            "targetGuid": target["guid"],
                            "queryScopeGuid": scope["guid"],
                            "queryScopePath": scope["path"],
                        }
                    )
            finally:
                progress.update(detail=f"{item.get('mac', '')} -> {item.get('targetPath', '')}")
    except Exception:
        progress.close("中止")
        raise
    progress.finish()
    return resolved


def dedupe_and_find_conflicts(resolved_moves: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    by_terminal: dict[str, dict[str, Any]] = {}
    for move in resolved_moves:
        existing = by_terminal.get(move["terminalGuid"])
        if existing is None:
            by_terminal[move["terminalGuid"]] = move
            continue
        if existing["targetGuid"] != move["targetGuid"]:
            existing["conflict"] = True
            move["conflict"] = True
    unique = list(by_terminal.values())
    conflicts = [move for move in unique if move.get("conflict")]
    movable = [move for move in unique if not move.get("conflict")]
    return unique, movable, conflicts


def verification_base(move: dict[str, Any]) -> dict[str, Any]:
    return {
        "action": "verify_move",
        "sourceRow": move.get("sourceRow"),
        "mac": move.get("mac"),
        "apiMac": move.get("apiMac") or format_mac_for_api(move.get("mac")),
        "terminalGuid": move.get("terminalGuid"),
        "terminalName": move.get("terminalName"),
        "hostName": move.get("hostName"),
        "targetPath": move.get("targetPath"),
        "targetGuid": move.get("targetGuid"),
        "queryScopePath": move.get("queryScopePath"),
        "queryScopeGuid": move.get("queryScopeGuid"),
    }


def verify_move_effect(client: ApiClient, plan: dict[str, Any], move: dict[str, Any]) -> dict[str, Any]:
    attempts = plan["config"].get("moveVerifyRetries", 0) + 1
    delay_seconds = plan["config"].get("moveVerifyDelaySeconds", 0.0)
    last_record: dict[str, Any] = {}

    for attempt in range(1, attempts + 1):
        record = verification_base(move)
        record["attempt"] = attempt
        record["maxAttempts"] = attempts
        try:
            matches = client.query_terminals_by_mac(
                text(move.get("queryScopeGuid")),
                text(move.get("apiMac")) or format_mac_for_api(move.get("mac")),
                plan["config"]["terminalQueryPageSize"],
            )
        except Exception as exc:
            record["status"] = "verify_failed"
            record["error"] = str(exc)
            return record

        matched_guids = [text(row.get("guid")) for row in matches if text(row.get("guid"))]
        record["matchCount"] = len(matches)
        record["matchedGuids"] = matched_guids
        for terminal in matches:
            if text(terminal.get("guid")) != text(move.get("terminalGuid")):
                continue
            current_group_guid = text(terminal.get("groupGuid"))
            record["currentGroupGuid"] = current_group_guid
            record["status"] = "verified" if current_group_guid == text(move.get("targetGuid")) else "not_effective"
            break
        else:
            record["status"] = "not_found_after_move"

        last_record = record
        if record["status"] == "verified":
            return record
        if attempt < attempts and delay_seconds:
            time.sleep(delay_seconds)

    return last_record


def move_verification_message(verification: dict[str, Any]) -> str:
    status = text(verification.get("status"))
    if status == "not_effective":
        return "终端移动请求已返回成功，但复查发现终端仍不在目标分组。"
    if status == "not_found_after_move":
        return "终端移动请求已返回成功，但复查时按 MAC 未找到该终端 GUID。"
    if status == "verify_failed":
        return f"终端移动请求已返回成功，但复查接口失败: {verification.get('error', '')}"
    return f"终端移动复查未通过: {status}"


def move_terminals(client: ApiClient, plan: dict[str, Any], execute: bool, resolved_moves: list[dict[str, Any]], report: dict[str, Any], progress_enabled: bool = True) -> None:
    unique, movable, conflicts = dedupe_and_find_conflicts(resolved_moves)
    report["moves"]["resolvedTerminalGuidCount"] = len(unique)
    report["moves"]["conflicts"] = conflicts
    report["moves"].setdefault("verifications", [])

    for conflict in conflicts:
        report["validation"].append({"severity": "error", "source": "终端清单", "field": "MAC地址", "value": conflict["mac"], "message": f"同一终端 GUID 在本次计划中指向多个目标分组，已跳过: {conflict['terminalGuid']}"})

    progress = ProgressBar(len(movable), f"[{report.get('networkName', '')}] 移动复查", progress_enabled)
    try:
        for move in movable:
            record = {
                "action": "move_terminals",
                "status": "pending" if execute and plan["config"]["moveTerminals"] else "dry_run",
                "requestMode": "single_terminal",
                "batchSize": plan["config"]["executeBatchSize"],
                "sourceRow": move.get("sourceRow"),
                "mac": move.get("mac"),
                "apiMac": move.get("apiMac"),
                "terminalName": move.get("terminalName"),
                "hostName": move.get("hostName"),
                "currentGroupGuid": move.get("currentGroupGuid"),
                "targetPath": move["targetPath"],
                "targetGuid": move["targetGuid"],
                "terminalGuidList": [move["terminalGuid"]],
            }
            try:
                if text(move.get("currentGroupGuid")) == text(move.get("targetGuid")):
                    record["status"] = "already_in_target"
                    verification = verification_base(move)
                    verification["status"] = "already_in_target"
                    verification["currentGroupGuid"] = move.get("currentGroupGuid")
                    report["moves"]["batches"].append(record)
                    report["moves"]["verifications"].append(verification)
                    continue

                if not execute:
                    record["status"] = "dry_run"
                    report["moves"]["batches"].append(record)
                    continue

                if not plan["config"]["moveTerminals"]:
                    record["status"] = "move_disabled"
                    report["moves"]["batches"].append(record)
                    continue

                try:
                    response = client.move_terminals([move["terminalGuid"]], move["targetGuid"])
                except Exception as exc:
                    record["status"] = "move_failed"
                    record["error"] = str(exc)
                    report["moves"]["batches"].append(record)
                    report["validation"].append(
                        {
                            "severity": "error",
                            "source": "终端移动",
                            "row": move.get("sourceRow"),
                            "field": "terminalGuid",
                            "value": move.get("terminalGuid"),
                            "message": f"终端移动接口失败: {exc}",
                        }
                    )
                    continue

                record["status"] = "moved"
                record["response"] = response
                verification = verify_move_effect(client, plan, move)
                record["verifyStatus"] = verification.get("status")
                report["moves"]["batches"].append(record)
                report["moves"]["verifications"].append(verification)
                if verification.get("status") != "verified":
                    report["validation"].append(
                        {
                            "severity": "error",
                            "source": "终端移动复查",
                            "row": move.get("sourceRow"),
                            "field": "terminalGuid",
                            "value": move.get("terminalGuid"),
                            "message": move_verification_message(verification),
                        }
                    )
            finally:
                progress.update(detail=f"{move.get('mac', '')} {move.get('terminalName') or move.get('hostName') or move.get('terminalGuid')}")
    except Exception:
        progress.close("中止")
        raise
    progress.finish()


def status_counts(items: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for item in items:
        status = text(item.get("status")) or "unknown"
        counts[status] = counts.get(status, 0) + 1
    return counts


def format_status_counts(counts: dict[str, int]) -> str:
    if not counts:
        return "0"
    return "，".join(f"{key} {value}" for key, value in counts.items())


def build_text_summary(report: dict[str, Any]) -> str:
    counts = report.get("templateCounts") or {}
    by_network = counts.get("plannedTerminalMovesByNetwork") or {}
    validation = report.get("validation") or []
    errors = [item for item in validation if item.get("severity") == "error"]
    warnings = [item for item in validation if item.get("severity") == "warning"]
    lines = [
        "资产分组运行摘要",
        "=" * 60,
        f"状态: {report.get('status', '')}",
        f"模式: {report.get('mode', '')}",
        f"开始: {report.get('startedAt', '')}",
        f"完成: {report.get('completedAt', '')}",
        f"模板: {report.get('templatePath', '')}",
        (
            "模板计划: "
            f"启用分组行 {counts.get('enabledGroupRows', 0)} 行，"
            f"本次未使用分组行 {counts.get('unusedEnabledGroupRows', 0)} 行，"
            f"目标分组 {counts.get('plannedTargetGroups', 0)} 项，"
            f"路径保障 {counts.get('plannedGroups', 0)} 项，"
            f"终端 {counts.get('plannedTerminalMoves', 0)} 项"
        ),
    ]
    if by_network:
        network_parts = [
            f"{network_label(key)} {value} 项"
            for key, value in by_network.items()
            if value
        ]
        lines.append(f"终端网络分布: {'，'.join(network_parts) if network_parts else '无'}")
    lines.append(f"错误: {len(errors)}，警告: {len(warnings)}")

    platform_runs = report.get("platformRuns") or []
    if platform_runs:
        lines.extend(["", "平台执行"])
        lines.append("-" * 60)
    for run in platform_runs:
        groups = run.get("groups") or []
        terminal_queries = run.get("terminalQueries") or []
        moves = run.get("moves") or {}
        batches = moves.get("batches") or []
        verifications = moves.get("verifications") or []
        matched = sum(1 for item in terminal_queries if item.get("status") == "matched")
        not_found = sum(1 for item in terminal_queries if item.get("status") == "not_found")
        lines.extend(
            [
                f"[{run.get('networkName')}] {run.get('status')} | {run.get('apiBaseUrl')}",
                f"  平台分组总数: {(run.get('platform') or {}).get('groupCount', 0)}",
                f"  分组动作: {format_status_counts(status_counts(groups))}",
                f"  终端查询: matched {matched}，not_found {not_found}，total {len(terminal_queries)}",
                f"  移动任务: {format_status_counts(status_counts(batches))}",
                f"  移动复查: {format_status_counts(status_counts(verifications))}",
                f"  已解析终端 GUID: {moves.get('resolvedTerminalGuidCount', 0)}",
            ]
        )
        for item in terminal_queries[:8]:
            lines.append(
                "    - "
                f"row {item.get('sourceRow')}: {item.get('mac')} -> {item.get('targetPath')} "
                f"({item.get('status')}, matches={item.get('matchCount', 0)})"
            )
        if len(terminal_queries) > 8:
            lines.append(f"    - 其余 {len(terminal_queries) - 8} 条终端查询见 JSON 明细")

    if errors or warnings:
        lines.extend(["", "校验与告警"])
        lines.append("-" * 60)
        for item in (errors + warnings)[:20]:
            prefix = item.get("networkName")
            network_part = f"[{prefix}] " if prefix else ""
            row_part = f" row {item.get('row')}" if item.get("row") else ""
            lines.append(f"- {network_part}{item.get('severity')}{row_part} {item.get('source', '')}/{item.get('field', '')}: {item.get('message', '')}")
        if len(errors) + len(warnings) > 20:
            lines.append(f"- 其余 {len(errors) + len(warnings) - 20} 条见 JSON 明细")

    return "\n".join(lines) + "\n"


def format_validation_line(item: dict[str, Any]) -> str:
    network_part = f"[{item.get('networkName')}] " if item.get("networkName") else ""
    row_part = f" row {item.get('row')}" if item.get("row") else ""
    source = item.get("source", "")
    field = item.get("field", "")
    message = item.get("message", "")
    return f"{network_part}{item.get('severity', '')}{row_part} {source}/{field}: {message}"


def print_validation_summary(validation: list[dict[str, Any]], limit: int = 10) -> None:
    errors = [item for item in validation if item.get("severity") == "error"]
    warnings = [item for item in validation if item.get("severity") == "warning"]
    if not errors and not warnings:
        return
    print("\n错误/警告明细（前 10 条）")
    print("-" * 60)
    for item in (errors + warnings)[:limit]:
        print(f"- {format_validation_line(item)}")
    remaining = len(errors) + len(warnings) - limit
    if remaining > 0:
        print(f"- 其余 {remaining} 条见 JSON 报告")


def write_report(report: dict[str, Any], output_dir: Path, keep_history: bool = False) -> tuple[Path | None, Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    latest_path = output_dir / "asset-grouping-run-latest.json"
    summary_latest_path = output_dir / "asset-grouping-summary-latest.txt"
    content = json.dumps(report, ensure_ascii=False, indent=2)
    summary_content = build_text_summary(report)
    latest_path.write_text(content, encoding="utf-8")
    summary_latest_path.write_text(summary_content, encoding="utf-8")
    report_path = None
    if keep_history:
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        report_path = output_dir / f"asset-grouping-run-{stamp}.json"
        summary_path = output_dir / f"asset-grouping-summary-{stamp}.txt"
        report_path.write_text(content, encoding="utf-8")
        summary_path.write_text(summary_content, encoding="utf-8")
    return report_path, latest_path, summary_latest_path


def mask_secret(value: str, keep: int = 6) -> str:
    value = text(value)
    if not value:
        return "(空)"
    if len(value) <= keep * 2:
        return "*" * len(value)
    return f"{value[:keep]}...{value[-keep:]}"


def print_report_summary(report_path: Path) -> None:
    if not report_path.exists():
        print(f"暂无报告: {report_path}")
        return
    report = json.loads(report_path.read_text(encoding="utf-8"))
    summary_path = report_path.with_name("asset-grouping-summary-latest.txt")
    print("\n最近报告")
    print("-" * 60)
    if summary_path.exists():
        print(summary_path.read_text(encoding="utf-8").rstrip())
    else:
        print(build_text_summary(report).rstrip())
    print(f"JSON 明细: {report_path}")
    print(f"摘要文本: {summary_path}")


def scan_invalid_macs(template_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(template_path, read_only=True, data_only=True)
    if "终端清单" not in workbook.sheetnames:
        raise RuntimeError(f"模板缺少工作表: 终端清单 ({template_path})")
    ws = workbook["终端清单"]
    invalid_rows: list[dict[str, Any]] = []
    for row in row_dicts(ws, 6, ws.max_row, ws.max_column):
        if not enabled(row.get("启用")):
            continue
        raw_mac = text(row.get("MAC地址"))
        normalized = normalize_mac(raw_mac)
        if valid_normalized_mac(normalized):
            continue
        reasons: list[str] = []
        if not raw_mac:
            reasons.append("MAC为空")
        if normalized and len(normalized) != 12:
            reasons.append(f"规范化后长度为 {len(normalized)}，不是 12")
        bad_chars = invalid_mac_chars(normalized)
        if bad_chars:
            reasons.append(f"包含非十六进制字符: {bad_chars}")
        if not reasons:
            reasons.append("无法规范化为 12 位十六进制 MAC")
        invalid_rows.append(
            {
                "sourceRow": row["__rowNumber"],
                "assetCode": text(row.get("资产编号")),
                "mac": raw_mac,
                "normalizedMac": normalized,
                "invalidChars": bad_chars,
                "reason": "；".join(reasons),
                "dept": text(row.get("设备责任科室")),
                "networkArea": text(row.get("网络区域")),
                "campus": text(row.get("院区")),
                "building": text(row.get("楼栋")),
                "floor": text(row.get("楼层")),
                "area": text(row.get("分区")),
                "location": text(row.get("具体物理地址")),
                "manualTargetPath": text(row.get("手动目标分组路径")),
                "remark": text(row.get("备注")),
            }
        )
    return invalid_rows


def export_invalid_macs(args: argparse.Namespace, output_path: Path | None = None) -> int:
    template_path = Path(args.template)
    output_dir = Path(args.output_dir)
    invalid_rows = scan_invalid_macs(template_path)
    output_path = output_path or output_dir / "invalid-mac-latest.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    ws = workbook.active
    ws.title = "不合法MAC"
    headers = [
        "源行",
        "资产编号",
        "原MAC",
        "规范化后",
        "非法字符",
        "问题",
        "设备责任科室",
        "网络区域",
        "院区",
        "楼栋",
        "楼层",
        "分区",
        "具体物理地址",
        "手动目标分组路径",
        "备注",
    ]
    ws.append(headers)
    for item in invalid_rows:
        ws.append(
            [
                item["sourceRow"],
                item["assetCode"],
                item["mac"],
                item["normalizedMac"],
                item["invalidChars"],
                item["reason"],
                item["dept"],
                item["networkArea"],
                item["campus"],
                item["building"],
                item["floor"],
                item["area"],
                item["location"],
                item["manualTargetPath"],
                item["remark"],
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    warning_fill = PatternFill("solid", fgColor="FCE4D6")
    white_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = warning_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        column = get_column_letter(col)
        max_len = max(len(str(cell.value or "")) for cell in ws[column])
        ws.column_dimensions[column].width = min(max(max_len + 2, 10), 55)

    workbook.save(output_path)
    print("\n不合法 MAC 导出完成")
    print("-" * 60)
    print(f"模板: {template_path}")
    print(f"数量: {len(invalid_rows)}")
    print(f"路径: {output_path}")
    if invalid_rows:
        print("前 10 条:")
        for item in invalid_rows[:10]:
            print(f"- 行 {item['sourceRow']} | {item['assetCode']} | {item['mac']} | {item['reason']}")
    return 0


def collect_not_found_terminals(report: dict[str, Any]) -> list[dict[str, Any]]:
    move_by_row: dict[int, dict[str, Any]] = {}
    for item in ((report.get("localPlan") or {}).get("movePlan") or []):
        source_row = item.get("sourceRow")
        if isinstance(source_row, int):
            move_by_row[source_row] = item

    rows: list[dict[str, Any]] = []
    for query in report.get("terminalQueries") or []:
        if query.get("status") != "not_found":
            continue
        source_row = query.get("sourceRow")
        plan_item = move_by_row.get(source_row) if isinstance(source_row, int) else None
        rows.append(
            {
                "sourceRow": source_row,
                "networkName": query.get("networkName") or (plan_item or {}).get("networkName"),
                "assetCode": (plan_item or {}).get("assetCode", ""),
                "mac": query.get("mac") or (plan_item or {}).get("mac", ""),
                "apiMac": query.get("apiMac") or (plan_item or {}).get("apiMac", ""),
                "hostname": (plan_item or {}).get("hostname", ""),
                "dept": (plan_item or {}).get("dept", ""),
                "targetPath": query.get("targetPath") or (plan_item or {}).get("targetPath", ""),
                "targetSource": (plan_item or {}).get("targetSource", ""),
                "matchCount": query.get("matchCount", 0),
                "status": query.get("status", ""),
                "location": (plan_item or {}).get("location", ""),
                "mappingSourceRow": (plan_item or {}).get("mappingSourceRow", ""),
                "queryScopePath": query.get("queryScopePath", ""),
                "message": "平台按 MAC 未查询到终端",
            }
        )
    return rows


def export_not_found_terminals(args: argparse.Namespace, output_path: Path | None = None) -> int:
    output_dir = Path(args.output_dir)
    report_path = output_dir / "asset-grouping-run-latest.json"
    if not report_path.exists():
        print(f"未找到最近报告: {report_path}")
        return 2

    report = json.loads(report_path.read_text(encoding="utf-8"))
    rows = collect_not_found_terminals(report)
    output_path = output_path or output_dir / "not-found-terminals-latest.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    ws = workbook.active
    ws.title = "未查询到终端"
    headers = [
        "源行",
        "网络",
        "资产编号",
        "原MAC",
        "查询MAC",
        "主机名",
        "设备责任科室",
        "目标分组路径",
        "目标来源",
        "匹配数",
        "状态",
        "位置",
        "映射源行",
        "查询范围",
        "说明",
    ]
    ws.append(headers)
    for item in rows:
        ws.append(
            [
                item["sourceRow"],
                item["networkName"],
                item["assetCode"],
                item["mac"],
                item["apiMac"],
                item["hostname"],
                item["dept"],
                item["targetPath"],
                item["targetSource"],
                item["matchCount"],
                item["status"],
                item["location"],
                item["mappingSourceRow"],
                item["queryScopePath"],
                item["message"],
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    white_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = warning_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        column = get_column_letter(col)
        max_len = max(len(str(cell.value or "")) for cell in ws[column])
        ws.column_dimensions[column].width = min(max(max_len + 2, 10), 65)

    workbook.save(output_path)
    print("\n未查询到终端导出完成")
    print("-" * 60)
    print(f"报告: {report_path}")
    print(f"数量: {len(rows)}")
    print(f"路径: {output_path}")
    if not report.get("terminalQueries"):
        print("提示: 最近报告没有平台终端查询结果。请先执行平台 dry-run 或正式执行后再导出。")
    elif rows:
        print("前 10 条:")
        for item in rows[:10]:
            print(f"- 行 {item['sourceRow']} | {item['networkName']} | {item['assetCode']} | {item['mac']} -> {item['targetPath']}")
    return 0


def load_export_group_rows(export_path: Path) -> dict[str, dict[str, Any]]:
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Workbook contains no default style.*")
        workbook = load_workbook(export_path, read_only=True, data_only=True)
    ws = workbook[workbook.sheetnames[0]]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [text(value) for value in header]
    try:
        name_index = headers.index("分组名称")
        path_index = headers.index("分组路径")
    except ValueError as exc:
        raise RuntimeError("导出分组 Excel 必须包含表头：分组名称、分组路径") from exc
    desc_index = headers.index("分组描述") if "分组描述" in headers else None

    rows: dict[str, dict[str, Any]] = {}
    for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if path_index >= len(values):
            continue
        path = normalize_path(values[path_index])
        if not path:
            continue
        rows[path] = {
            "row": excel_row,
            "name": text(values[name_index]) if name_index < len(values) else "",
            "path": path,
            "description": text(values[desc_index]) if desc_index is not None and desc_index < len(values) else "",
        }
    return rows


def compare_export_groups(args: argparse.Namespace, export_path: Path) -> int:
    cookie = text(getattr(args, "cookie", "")) or text(getattr(args, "internal_cookie", ""))
    base_url = text(getattr(args, "base_url", "")) or text(getattr(args, "internal_base_url", ""))
    if not cookie:
        print("未配置 Cookie，无法查询平台实时分组。请先在配置中设置 Cookie，或使用 UES_COOKIE 环境变量。")
        return 2

    client = ApiClient(base_url, cookie, insecure=not args.strict_tls)
    group_tree = client.get_groups()
    platform_index = build_path_index(group_tree, ">")
    platform_by_path = {item["path"]: item for item in platform_index["flattened"]}
    export_by_path = load_export_group_rows(export_path)

    platform_paths = set(platform_by_path)
    export_paths = set(export_by_path)
    only_platform = sorted(platform_paths - export_paths)
    only_export = sorted(export_paths - platform_paths)
    common = sorted(platform_paths & export_paths)
    name_mismatch = [
        {
            "path": path,
            "platformName": platform_by_path[path]["name"],
            "exportName": export_by_path[path]["name"],
        }
        for path in common
        if platform_by_path[path]["name"] != export_by_path[path]["name"]
    ]

    report = {
        "createdAt": now_iso(),
        "exportPath": str(export_path),
        "platformCount": len(platform_paths),
        "exportCount": len(export_paths),
        "commonCount": len(common),
        "onlyPlatformCount": len(only_platform),
        "onlyExportCount": len(only_export),
        "nameMismatchCount": len(name_mismatch),
        "onlyPlatform": [
            {
                "path": path,
                "name": platform_by_path[path]["name"],
                "guid": platform_by_path[path]["guid"],
                "parentGuid": platform_by_path[path]["parentGuid"],
            }
            for path in only_platform
        ],
        "onlyExport": [export_by_path[path] for path in only_export],
        "nameMismatch": name_mismatch,
    }

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    latest_path = output_dir / "group-export-compare-latest.json"
    latest_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print("\n分组导出比对结果")
    print("-" * 60)
    print(f"平台实时分组: {len(platform_paths)}")
    print(f"导出 Excel 分组: {len(export_paths)}")
    print(f"两边一致路径: {len(common)}")
    print(f"平台有、导出没有: {len(only_platform)}")
    print(f"导出有、平台没有: {len(only_export)}")
    print(f"名称不一致: {len(name_mismatch)}")
    if only_platform:
        print("\n平台有、导出没有（前 20 条）:")
        for path in only_platform[:20]:
            item = platform_by_path[path]
            print(f"- {path} | guid={item['guid']}")
    if only_export:
        print("\n导出有、平台没有（前 20 条）:")
        for path in only_export[:20]:
            print(f"- {path}")
    print(f"\n比对报告: {latest_path}")
    return 0


def print_config(config: dict[str, Any], config_path: Path) -> None:
    print("\n当前配置")
    print("-" * 60)
    print(f"配置文件: {config_path}")
    print(f"template_path: {config.get('template_path')}")
    print(f"output_dir: {config.get('output_dir')}")
    print(f"api_base_url(兼容旧模板): {config.get('api_base_url')}")
    print(f"internal_api_base_url: {config.get('internal_api_base_url') or config.get('api_base_url')}")
    print(f"external_api_base_url: {config.get('external_api_base_url')}")
    print(f"query_scope_path: {config.get('query_scope_path')}")
    print(f"default_network: {config.get('default_network') or '内网'}")
    print(f"internal_cookie: {mask_secret(text(config.get('internal_cookie')))}")
    print(f"external_cookie: {mask_secret(text(config.get('external_cookie')))}")
    print(f"strict_tls: {bool(config.get('strict_tls'))}")
    print(f"keep_history: {bool(config.get('keep_history'))}")


def edit_config_menu(config_path: Path, config: dict[str, Any]) -> dict[str, Any]:
    while True:
        print_config(config, config_path)
        print("\n配置菜单")
        print("1. 设置内网 Cookie")
        print("2. 设置外网 Cookie")
        print("3. 清空内网 Cookie")
        print("4. 清空外网 Cookie")
        print("5. 清空全部 Cookie")
        print("6. 设置内网 API 地址")
        print("7. 设置外网 API 地址")
        print("8. 设置模板路径")
        print("9. 设置查询范围分组路径")
        print("10. 设置默认网络")
        print("11. 切换 strict_tls")
        print("12. 切换 keep_history")
        print("0. 返回")
        choice = input("请选择: ").strip()
        if choice == "0":
            return config
        if choice == "1":
            cookie = getpass.getpass("粘贴内网 Cookie（输入时不显示）: ").strip()
            config["internal_cookie"] = cookie
        elif choice == "2":
            cookie = getpass.getpass("粘贴外网 Cookie（输入时不显示）: ").strip()
            config["external_cookie"] = cookie
        elif choice == "3":
            config["internal_cookie"] = ""
        elif choice == "4":
            config["external_cookie"] = ""
        elif choice == "5":
            config["internal_cookie"] = ""
            config["external_cookie"] = ""
        elif choice == "6":
            value = input("内网 API 地址: ").strip()
            if value:
                config["internal_api_base_url"] = value
                config["api_base_url"] = value
        elif choice == "7":
            value = input("外网 API 地址: ").strip()
            if value:
                config["external_api_base_url"] = value
        elif choice == "8":
            value = input("模板路径: ").strip()
            if value:
                config["template_path"] = value
        elif choice == "9":
            value = input("查询范围分组路径: ").strip()
            if value:
                config["query_scope_path"] = value
        elif choice == "10":
            value = input("默认网络（内网/外网）: ").strip()
            if value:
                config["default_network"] = network_label(normalize_network_key(value, "internal"))
        elif choice == "11":
            config["strict_tls"] = not bool(config.get("strict_tls"))
        elif choice == "12":
            config["keep_history"] = not bool(config.get("keep_history"))
        else:
            print("无效选项。")
            continue
        save_config(config_path, config)
        print("配置已保存。")


def refresh_args_from_config(args: argparse.Namespace, config: dict[str, Any]) -> argparse.Namespace:
    env_cookie = os.getenv("UES_COOKIE") or os.getenv("OFFICE_SHIELD_COOKIE") or ""
    env_internal_cookie = os.getenv("UES_INTERNAL_COOKIE") or os.getenv("OFFICE_SHIELD_INTERNAL_COOKIE") or env_cookie
    env_external_cookie = os.getenv("UES_EXTERNAL_COOKIE") or os.getenv("OFFICE_SHIELD_EXTERNAL_COOKIE") or env_cookie
    common_cookie = text(config.get("cookie"))
    internal_cookie = text(config.get("internal_cookie")) or common_cookie
    external_cookie = text(config.get("external_cookie")) or common_cookie
    legacy_base_url = text(config.get("api_base_url")) or DEFAULT_CONFIG_DATA["api_base_url"]
    args.template = text(config.get("template_path")) or str(DEFAULT_TEMPLATE)
    args.output_dir = text(config.get("output_dir")) or str(DEFAULT_WORKDIR)
    args.cookie = env_cookie or common_cookie
    args.internal_cookie = env_internal_cookie or internal_cookie
    args.external_cookie = env_external_cookie or external_cookie
    args.internal_base_url = text(config.get("internal_api_base_url")) or legacy_base_url
    args.external_base_url = text(config.get("external_api_base_url")) or DEFAULT_CONFIG_DATA["external_api_base_url"]
    args.base_url = args.internal_base_url
    args.default_network = text(config.get("default_network")) or "内网"
    args.query_scope_path = text(config.get("query_scope_path")) or "全部终端"
    args.strict_tls = bool(config.get("strict_tls"))
    args.keep_history = bool(config.get("keep_history"))
    return args


def run_once(args: argparse.Namespace) -> int:
    template_path = Path(args.template)
    output_dir = Path(args.output_dir)
    default_network_key = normalize_network_key(getattr(args, "default_network", "内网"), "internal")
    plan = load_template_plan(template_path, default_network_key)

    execute = bool(args.execute and not plan["config"]["dryRun"])
    platform_configs = platform_configs_from_args(args)
    network_keys = planned_network_keys(plan)
    api_enabled = bool(
        not args.local_only
        and network_keys
        and any(platform_configs.get(key, {}).get("cookie") and platform_configs.get(key, {}).get("apiBaseUrl") for key in network_keys)
    )
    report: dict[str, Any] = {
        "startedAt": now_iso(),
        "mode": "execute" if api_enabled and execute else "platform_dry_run" if api_enabled else "local_only",
        "templatePath": str(template_path),
        "config": {
            "platforms": {
                key: {
                    "networkName": platform_configs[key]["networkName"],
                    "apiBaseUrl": platform_configs[key]["apiBaseUrl"],
                    "hasCookie": bool(platform_configs[key]["cookie"]),
                }
                for key in NETWORK_ORDER
            },
            "queryScopePath": args.query_scope_path,
            "defaultNetwork": network_label(default_network_key),
            "createMissingGroups": plan["config"]["createMissingGroups"],
            "moveTerminals": plan["config"]["moveTerminals"],
            "deleteExtraGroups": plan["config"]["deleteExtraGroups"],
            "batchSizeFromTemplate": plan["config"]["batchSize"],
            "executeBatchSize": plan["config"]["executeBatchSize"],
            "moveVerifyRetries": plan["config"]["moveVerifyRetries"],
            "moveVerifyDelaySeconds": plan["config"]["moveVerifyDelaySeconds"],
            "createDeclaredGroupsWithoutTerminals": plan["config"]["createDeclaredGroupsWithoutTerminals"],
            "multiTerminalMatchPolicy": plan["config"]["multiTerminalMatchPolicy"],
            "insecureTls": not args.strict_tls,
            "templateDryRun": plan["config"]["dryRun"],
            "commandExecute": bool(args.execute),
        },
        "templateCounts": plan["counts"],
        "validation": list(plan["validation"]),
        "localPlan": {
            "createGroupPlan": plan["createGroupPlan"],
            "targetGroupUsage": plan["targetGroupUsage"],
            "movePlan": plan["movePlan"],
        },
        "platform": {},
        "platformRuns": [],
        "groups": [],
        "terminalQueries": [],
        "moves": {"resolvedTerminalGuidCount": 0, "conflicts": [], "batches": [], "verifications": []},
    }

    if plan["counts"]["validationErrors"] > 0:
        report["completedAt"] = now_iso()
        report["status"] = "blocked_template_validation"
        _, latest_path, summary_path = write_report(report, output_dir, args.keep_history)
        print(f"模板校验失败，已停止。报告: {latest_path}；摘要: {summary_path}")
        print_validation_summary(report["validation"])
        return 2

    if not api_enabled:
        if not args.local_only:
            report["validation"].append({"severity": "warning", "source": "运行环境", "field": "Cookie", "value": "", "message": "未提供可用平台 Cookie，已仅生成本地计划。可在配置中填写内网/外网 Cookie，或使用 UES_INTERNAL_COOKIE / UES_EXTERNAL_COOKIE 环境变量。"})
        report["completedAt"] = now_iso()
        report["status"] = "local_plan_ready"
        _, latest_path, summary_path = write_report(report, output_dir, args.keep_history)
        print(
            "本地计划完成："
            f"启用分组行 {plan['counts']['enabledGroupRows']} 行，"
            f"本次未使用分组行 {plan['counts'].get('unusedEnabledGroupRows', 0)} 行，"
            f"目标分组 {plan['counts'].get('plannedTargetGroups', 0)} 项，"
            f"路径保障 {plan['counts']['plannedGroups']} 项，"
            f"终端 {plan['counts']['plannedTerminalMoves']} 项。"
            f"报告: {latest_path}；摘要: {summary_path}"
        )
        print_validation_summary(report["validation"])
        return 0

    if execute:
        missing_platforms = [
            platform_configs[key]["networkName"]
            for key in network_keys
            if not platform_configs.get(key, {}).get("apiBaseUrl") or not platform_configs.get(key, {}).get("cookie")
        ]
        if missing_platforms:
            report["validation"].append(
                {
                    "severity": "error",
                    "source": "运行环境",
                    "field": "platform_config",
                    "value": "，".join(missing_platforms),
                    "message": "正式执行前必须为所有涉及网络配置 API 地址和 Cookie。",
                }
            )
            report["completedAt"] = now_iso()
            report["status"] = "blocked_missing_platform_config"
            _, latest_path, summary_path = write_report(report, output_dir, args.keep_history)
            print(f"正式执行已停止，平台配置不完整。报告: {latest_path}；摘要: {summary_path}")
            print_validation_summary(report["validation"])
            return 2

    for network_key in network_keys:
        network_plan = filter_plan_for_network(plan, network_key)
        platform_config = platform_configs.get(network_key) or {
            "networkKey": network_key,
            "networkName": network_label(network_key),
            "apiBaseUrl": "",
            "cookie": "",
        }
        platform_report = new_platform_report(network_key, platform_config)
        if not network_plan["movePlan"]:
            platform_report["status"] = "skipped_no_terminal"
            merge_platform_report(report, platform_report)
            continue
        if not platform_config.get("apiBaseUrl") or not platform_config.get("cookie"):
            platform_report["status"] = "skipped_missing_config"
            platform_report["validation"].append(
                {
                    "severity": "error" if execute else "warning",
                    "source": "运行环境",
                    "field": "platform_config",
                    "value": platform_config["networkName"],
                    "message": f"{platform_config['networkName']} 平台缺少 API 地址或 Cookie，已跳过该网络。",
                }
            )
            print(f"\n跳过 {platform_config['networkName']}：缺少 API 地址或 Cookie")
            merge_platform_report(report, platform_report)
            continue
        try:
            client = ApiClient(platform_config["apiBaseUrl"], platform_config["cookie"], insecure=not args.strict_tls)
            print(f"\n开始处理 {platform_config['networkName']}：{platform_config['apiBaseUrl']}")
            group_index = ensure_groups(client, network_plan, execute, platform_report)
            if not any(item.get("severity") == "error" for item in platform_report["validation"]):
                resolved_moves = resolve_terminal_moves(client, network_plan, group_index, args.query_scope_path, platform_report)
                move_terminals(client, network_plan, execute, resolved_moves, platform_report)
            platform_report["status"] = "completed_with_errors" if any(item.get("severity") == "error" for item in platform_report["validation"]) else "executed" if execute else "platform_dry_run_ready"
        except Exception as exc:
            print(f"\n{platform_config['networkName']} 处理失败：{exc}")
            platform_report["status"] = "failed"
            platform_report["validation"].append(
                {
                    "severity": "error",
                    "source": "平台接口",
                    "field": platform_config["apiBaseUrl"],
                    "value": platform_config["networkName"],
                    "message": str(exc),
                }
            )
        merge_platform_report(report, platform_report)

    report["completedAt"] = now_iso()
    report["status"] = "completed_with_errors" if any(item.get("severity") == "error" for item in report["validation"]) else "executed" if execute else "platform_dry_run_ready"
    _, latest_path, summary_path = write_report(report, output_dir, args.keep_history)
    network_parts = [
        f"{network_label(key)} {plan['counts']['plannedTerminalMovesByNetwork'].get(key, 0)} 项"
        for key in NETWORK_ORDER
        if plan["counts"].get("plannedTerminalMovesByNetwork", {}).get(key, 0)
    ]
    print("\n".join([
        f"模式: {report['mode']}",
        f"状态: {report['status']}",
        (
            "模板计划: "
            f"启用分组行 {plan['counts']['enabledGroupRows']} 行，"
            f"本次未使用分组行 {plan['counts'].get('unusedEnabledGroupRows', 0)} 行，"
            f"目标分组 {plan['counts'].get('plannedTargetGroups', 0)} 项，"
            f"路径保障 {plan['counts']['plannedGroups']} 项，"
            f"终端 {plan['counts']['plannedTerminalMoves']} 项"
        ),
        f"终端网络分布: {'，'.join(network_parts) if network_parts else '无'}",
        f"平台分组动作: {len(report['groups'])} 项",
        f"终端查询: {len(report['terminalQueries'])} 项",
        f"移动任务: {len(report['moves']['batches'])} 项",
        f"移动复查: {format_status_counts(status_counts(report['moves'].get('verifications') or []))}",
        f"报告: {latest_path}",
        f"摘要: {summary_path}",
    ]))
    print_validation_summary(report["validation"])
    return 0


def run_interactive(args: argparse.Namespace, config_path: Path, config: dict[str, Any]) -> int:
    while True:
        print("\n资产分组工具")
        print("=" * 60)
        print("1. 本地解析模板（不访问平台）")
        print("2. 平台 dry-run（查平台，不新建/不移动）")
        print("3. 正式执行（需模板 dry_run=FALSE）")
        print("4. 查看/修改配置")
        print("5. 查看最近报告")
        print("6. 比对导出的分组 Excel")
        print("7. 导出不合法 MAC")
        print("8. 导出未查询到终端")
        print("0. 退出")
        choice = input("请选择: ").strip()

        if choice == "0":
            return 0
        if choice == "1":
            run_args = argparse.Namespace(**vars(args))
            run_args.local_only = True
            run_args.execute = False
            run_once(run_args)
        elif choice == "2":
            run_args = argparse.Namespace(**vars(args))
            run_args.local_only = False
            run_args.execute = False
            if not any(item.get("cookie") for item in platform_configs_from_args(run_args).values()):
                print("未配置 Cookie。请先在配置菜单中设置内网/外网 Cookie，或使用 UES_COOKIE 环境变量。")
                continue
            run_once(run_args)
        elif choice == "3":
            run_args = argparse.Namespace(**vars(args))
            run_args.local_only = False
            run_args.execute = True
            if not any(item.get("cookie") for item in platform_configs_from_args(run_args).values()):
                print("未配置 Cookie。请先在配置菜单中设置内网/外网 Cookie，或使用 UES_COOKIE 环境变量。")
                continue
            print("正式执行会调用平台接口新建分组并移动终端。")
            confirm = input("确认请输入 EXECUTE: ").strip()
            if confirm != "EXECUTE":
                print("已取消。")
                continue
            run_once(run_args)
        elif choice == "4":
            config = edit_config_menu(config_path, config)
            args = refresh_args_from_config(args, config)
        elif choice == "5":
            print_report_summary(Path(args.output_dir) / "asset-grouping-run-latest.json")
        elif choice == "6":
            value = input("请输入导出的分组 Excel 路径: ").strip().strip('"')
            if not value:
                print("未输入路径。")
                continue
            compare_export_groups(args, Path(value))
        elif choice == "7":
            export_invalid_macs(args)
        elif choice == "8":
            export_not_found_terminals(args)
        else:
            print("无效选项。")


def main() -> int:
    raw_args = parse_args()
    config_path = Path(raw_args.config)
    config = load_config(config_path)

    if raw_args.init_config:
        save_config(config_path, config)
        print(f"配置文件已创建/补全: {config_path}")
        return 0

    cli_requested_run = any(
        [
            raw_args.run,
            raw_args.local_only,
            raw_args.execute,
            raw_args.init_config,
            raw_args.template,
            raw_args.output_dir,
            raw_args.cookie,
            raw_args.base_url,
            raw_args.internal_cookie,
            raw_args.external_cookie,
            raw_args.internal_base_url,
            raw_args.external_base_url,
            raw_args.default_network,
            raw_args.query_scope_path,
            raw_args.compare_export,
            raw_args.export_invalid_mac is not None,
            raw_args.export_not_found_terminals is not None,
            raw_args.strict_tls,
            raw_args.keep_history,
        ]
    )
    args = apply_config(raw_args, config)
    if args.compare_export:
        return compare_export_groups(args, Path(args.compare_export))
    if args.export_invalid_mac is not None:
        output_path = Path(args.export_invalid_mac) if text(args.export_invalid_mac) else None
        return export_invalid_macs(args, output_path)
    if args.export_not_found_terminals is not None:
        output_path = Path(args.export_not_found_terminals) if text(args.export_not_found_terminals) else None
        return export_not_found_terminals(args, output_path)
    should_interactive = args.interactive or not cli_requested_run
    if should_interactive:
        return run_interactive(args, config_path, config)
    return run_once(args)


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("已中断。", file=sys.stderr)
        raise SystemExit(130)
    except Exception as exc:
        print(f"错误: {exc}", file=sys.stderr)
        raise SystemExit(1)
